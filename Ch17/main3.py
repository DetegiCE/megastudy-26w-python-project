import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re
import os

filePath = os.path.join(os.path.dirname(__file__), '2025년 고등교육기관 주소록(2025.4.1.)_250829my.xlsx')
df_from_excel = pd.read_excel(filePath, engine='openpyxl')
df_from_excel.columns = df_from_excel.loc[4].tolist()
df_from_excel = df_from_excel.drop(index=list(range(0,5)))

def request_geo(road):
    apiurl = "https://api.vworld.kr/req/address?"
    params = {
        "service": "address",
        "request": "getcoord",
        "crs": "epsg:4326",
        "address": road,
        "format": "json",
        "type": "road",
        "key": '발급받은 인증키 입력'
    }

    response = requests.get(apiurl, params=params)
    json_data = response.json()
    if json_data['response']['status'] == 'OK':
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x, y
    else:
        return 0,0

try:
    new_path = os.path.join(os.path.dirname(__file__), '학교주소좌표.xlsx')
    wb = load_workbook(new_path, data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active

university_list = df_from_excel['학교명'].to_list()
address_list = df_from_excel['주소'].to_list()

for num, value in enumerate(address_list):
    addr = re.sub(r'\([^)]*\)', '', value)
    print(addr)
    x, y = request_geo(addr)
    sheet.append([university_list[num], addr, x, y])

wb.save(new_path)
wb.close()
