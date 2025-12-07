from openpyxl import load_workbook
import os

file_path = os.path.join(os.path.dirname(__file__), '수료증명단.xlsx')
load_wb = load_workbook(file_path)
load_ws = load_wb.active

name_list = []
birthday_list = []
ho_list = []
for i in range(1, load_ws.max_row + 1):
    name = load_ws.cell(i, 1).value
    birthday = load_ws.cell(i, 2).value
    ho = load_ws.cell(i, 3).value

    name_list.append(name)
    birthday_list.append(birthday)
    ho_list.append(ho)

print(name_list)
print(birthday_list)
print(ho_list)
